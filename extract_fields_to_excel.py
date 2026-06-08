"""
extract_payroll_excel.py
------------------------
Extracts payroll data from a CSV (PDF text-extraction output) and writes
a formatted Excel file with columns:

  File Name | Name | Address Line 1 | Address Line 2 | Address Line 3 | Gross Pay | Net Pay

Usage:
    python extract_payroll_excel.py input.csv [output.xlsx]
"""

import csv, re, sys, time
from pathlib import Path

# ── Patterns ──────────────────────────────────────────────────────────────
TRIGGER_PATTERNS = [
    r"PURCHASE,?\s*NY\s*\d+",
    r"Pay\s*Date\s*:\s*\d+",
]

# Lines containing any of these → skip the entire row (bank/deposit stub junk)
SKIP_LINE_PATTERNS = [
    r"NON-NEGOTIABLE",
    r"Deposited to the account",
    r"account number transit",
    r"SPECIAL LEDGER",
    r"JWP\s+CNG",
    r"xxxx",                       # masked account numbers
    r"ABA\s+amount",
]

AMOUNT_PAT = re.compile(
    r'(-?\$)'
    r'('
    r'\d{1,3}(?:,\d{3})*(?:\.\d{2})'
    r'|\d{1,3}(?:\s\d{3})*\s\d{2}'
    r'|\d{1,3}\s\d{2}'
    r')'
)

# ── Progress Bar ───────────────────────────────────────────────────────────
class ProgressBar:
    def __init__(self, total, label="", width=45):
        self.total   = max(total, 1)
        self.label   = label
        self.width   = width
        self.current = 0
        self._draw(0)

    def _draw(self, n):
        pct    = n / self.total
        filled = int(self.width * pct)
        bar    = "█" * filled + "░" * (self.width - filled)
        text   = f"\r  {self.label:<18} [{bar}] {n:>6,}/{self.total:,}  {pct*100:5.1f}%"
        sys.stdout.write(text)
        sys.stdout.flush()

    def update(self, n=1):
        self.current += n
        self._draw(self.current)

    def done(self, msg=""):
        self._draw(self.total)
        print(f"  ✓ {msg}" if msg else "")


# ── Helpers ────────────────────────────────────────────────────────────────
def clean(text):
    return text.strip().strip('"').strip()

def is_trigger(text):
    return any(re.search(p, text, re.IGNORECASE) for p in TRIGGER_PATTERNS)

def is_skip_line(text):
    """Returns True for bank deposit / check stub lines — never capture these."""
    return any(re.search(p, text, re.IGNORECASE) for p in SKIP_LINE_PATTERNS)

def extract_name(text):
    after = re.split(r':', text)[-1].strip()
    after = re.sub(
        r'^(Married|Single|Head\s+of\s+Household|Qualifying\s+Widow\w*)\s*',
        '', after, flags=re.IGNORECASE
    ).strip()
    name_words = []
    for w in after.split():
        if re.match(r"^[A-Za-z][A-Za-z'.,-]*$", w):
            name_words.append(w)
        else:
            break
    return " ".join(name_words) if len(name_words) >= 2 else None

def is_address(text):
    """Detect if this line contains address info (street, apt, city/state/zip, etc.)"""
    # Skip lines like "NY: 1" or "NY: 0,$3 Additional Tax" (field lines, not address)
    if re.match(r'^[A-Z]{2}:\s*[\d,$\s]', text):
        return False
    
    # Street address: digits + words
    if re.match(r'^\d+\s+\w', text):
        return True
    # City/State/ZIP: Word(s) State ZIP
    if re.match(r'^[A-Z][A-Z\s]+\s+[A-Z]{2}\s+\d{5}', text):
        return True
    # Apartment/Unit: "APT", "UNIT", "STE", etc.
    if re.search(r'(?:APT|UNIT|STE|SUITE|FLOOR)\s*#?\d', text, re.IGNORECASE):
        return True
    # Exemptions/Allowances or Federal prefix
    if re.search(r'(?:Exemptions|Allowances|Federal):', text, re.IGNORECASE):
        return True
    return False

def clean_address(text):
    """Extract just the address part, stripping prefixes and metadata."""
    # Exemptions/Allowances: 155 FERRIS AVE
    m = re.search(r'(?:Exemptions|Allowances)[^:]*:\s*(.+)$', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    
    # Federal: 0 APT #7L  or  Federal: 0 DANBURY CT 06810
    m = re.search(r'Federal:\s*\d+\s*(.+)$', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    
    # Already plain address, return as-is
    return text.strip()

def parse_amount(sign_str, num_str):
    sign = -1 if '-' in sign_str else 1
    if '.' in num_str:
        return float(num_str.replace(',', '')) * sign
    d = num_str.replace(' ', '')
    return float(f"{d[:-2]}.{d[-2:]}") * sign

def first_amount_after(text, keyword):
    idx = text.lower().find(keyword.lower())
    if idx == -1: return None
    m = AMOUNT_PAT.search(text, idx + len(keyword))
    return parse_amount(m.group(1), m.group(2)) if m else None

def fmt(val):
    return f"${val:,.2f}" if val is not None else ""

def parse_address_parts(address_lines):
    """
    Split address lines into structured parts:
    - street_address (with unit/apt if present)
    - city_state_zip (formatted as "CITY, ST ZIP")
    
    Returns: (street_address, city_state_zip)
    """
    if not address_lines:
        return "", ""
    
    addr = address_lines + ["", "", ""]  # pad
    street = ""
    unit = ""
    city_state_zip = ""
    
    # Line 1: usually street address
    if addr[0]:
        street = addr[0].strip()
    
    # Line 2: could be unit/apt OR city/state/zip
    if addr[1]:
        if re.search(r'(?:APT|UNIT|STE|SUITE|FLOOR)\s*#?\d', addr[1], re.IGNORECASE):
            unit = addr[1].strip()
        elif re.match(r'^[A-Z][A-Z\s]+\s+[A-Z]{2}\s+\d{5}', addr[1]):
            city_state_zip = addr[1].strip()
    
    # Line 3: usually city/state/zip, or unit if line 2 was city
    if addr[2]:
        if not city_state_zip:
            city_state_zip = addr[2].strip()
    
    # Combine street + unit
    full_street = street
    if unit:
        full_street = f"{street} {unit}".strip() if street else unit
    
    # Format city, state, zip with comma
    if city_state_zip:
        # Parse "CITY STATE ZIP" → "CITY, ST ZIP"
        parts = city_state_zip.split()
        if len(parts) >= 3:
            # Last part is ZIP, second-to-last is STATE
            zip_code = parts[-1]
            state = parts[-2]
            city = " ".join(parts[:-2])
            city_state_zip = f"{city}, {state} {zip_code}"
    
    return full_street, city_state_zip


# ── Step 1: Load CSV ───────────────────────────────────────────────────────
def load_csv(csv_path):
    print("\n📂 Loading CSV...")
    raw = []
    with open(csv_path, newline='', encoding='utf-8') as fh:
        reader = csv.reader(fh)
        all_rows = list(reader)

    total = len(all_rows)
    pb = ProgressBar(total, label="Reading rows")
    for row in all_rows:
        pb.update()
        if len(row) < 3: continue
        fn, pg, tx = clean(row[0]), clean(row[1]), clean(row[2])
        if fn.lower() in ("file name", "filename"): continue
        if is_skip_line(tx): continue          # ← skip bank/deposit stub lines
        raw.append((fn, pg, tx))
    pb.done(f"{len(raw):,} data rows loaded")
    return raw


# ── Step 2: Extract records ────────────────────────────────────────────────
def extract_records(rows):
    print("\n🔍 Extracting records...")
    pb = ProgressBar(len(rows), label="Parsing rows")

    records, current = [], None
    state, addr_count = "search", 0

    for fn, pg, text in rows:
        pb.update()

        if is_trigger(text):
            if current and current["name"]: records.append(current)
            current = {"file_name": fn, "page": pg, "name": "",
                       "address": [], "gross_pay": None, "net_pay": None}
            state, addr_count = "name", 0
            continue

        if state == "name" and current is not None:
            name = extract_name(text)
            if name:
                current["name"] = name
                state, addr_count = "addr", 0
            continue

        if state == "addr" and current is not None:
            if is_address(text) and addr_count < 3:
                current["address"].append(clean_address(text))
                addr_count += 1
                continue
            # Skip metadata lines like "NY: 1", stay in address mode
            elif re.match(r'^[A-Z]{2}:\s*[\d,$\s]', text):
                continue
            elif addr_count > 0:
                state = "salary"

        if state == "salary" and current is not None:
            tl = text.lower()
            if "gross pay" in tl and current["gross_pay"] is None:
                current["gross_pay"] = first_amount_after(text, "Gross Pay")
            if "net pay" in tl and current["net_pay"] is None:
                current["net_pay"] = first_amount_after(text, "Net Pay")
            if "net check" in tl and current["net_pay"] is None:
                current["net_pay"] = first_amount_after(text, "Net Check")

    if current and current["name"]:
        records.append(current)

    pb.done(f"{len(records):,} employee records found")
    return records


# ── Step 3: Write Excel (separate file per source PDF) ────────────────────
def write_excel(records, output_dir="."):
    print("\n📊 Writing Excel files by source...")
    
    # Group records by source file
    files_data = {}
    for rec in records:
        fn = rec["file_name"]
        if fn not in files_data:
            files_data[fn] = []
        files_data[fn].append(rec)
    
    output_paths = []
    for file_name, file_records in files_data.items():
        # Generate output filename: Q00897.01-xxx.pdf → Q00897.01-xxx_payroll.xlsx
        base_name = file_name.rsplit('.', 1)[0]  # remove .pdf
        out_file = f"{output_dir}/{base_name}_payroll.xlsx"
        
        pb = ProgressBar(len(file_records), label=f"  {file_name}")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Payroll"

            HEADERS    = ["File Name", "Name", "Street Address", "City, State ZIP", "Gross Pay", "Net Pay"]
            COL_WIDTHS = [45,          25,     40,                 28,               15,          15]

            hdr_fill  = PatternFill("solid", fgColor="1F3864")
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            alt_fill  = PatternFill("solid", fgColor="EEF2FA")
            mono_font = Font(name="Courier New", size=10)
            thin      = Side(style="thin", color="D0D0D0")
            bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[cell.column_letter].width = w
            ws.row_dimensions[1].height = 24

            for ri, rec in enumerate(file_records, 2):
                pb.update()
                street, city_zip = parse_address_parts(rec["address"])
                row_vals = [
                    rec["file_name"], 
                    rec["name"],
                    street,
                    city_zip,
                    rec["gross_pay"], 
                    rec["net_pay"],
                ]
                fill = alt_fill if ri % 2 == 0 else None
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border    = bdr
                    cell.alignment = Alignment(vertical="center", wrap_text=(ci <= 4))
                    if fill: cell.fill = fill
                    if ci in (5, 6):  # Gross Pay, Net Pay columns
                        cell.number_format = '#,##0.00'
                        cell.font          = mono_font
                        cell.alignment     = Alignment(horizontal="right", vertical="center")
                ws.row_dimensions[ri].height = 18

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:F{len(file_records)+1}"

            wb.save(out_file)
            pb.done(f"✓ {len(file_records)} rows")
            output_paths.append((file_name, out_file, len(file_records)))
            
        except ImportError:
            pb.done()
            csv_out = out_file.replace(".xlsx", ".csv")
            with open(csv_out, "w", newline='', encoding='utf-8') as fh:
                w = csv.writer(fh)
                w.writerow(["File Name","Name","Street Address","City, State ZIP","Gross Pay","Net Pay"])
                for rec in file_records:
                    street, city_zip = parse_address_parts(rec["address"])
                    w.writerow([rec["file_name"], rec["name"],
                                 street, city_zip,
                                 fmt(rec["gross_pay"]), fmt(rec["net_pay"])])
            output_paths.append((file_name, csv_out, len(file_records)))
    
    return output_paths


# ── Entry point ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_payroll_excel.py <input.csv> [output_dir]")
        print("       Writes separate Excel files for each source PDF")
        sys.exit(1)

    in_csv  = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else "."

    if not Path(in_csv).exists():
        print(f"Error: {in_csv} not found")
        sys.exit(1)

    t0 = time.time()

    rows    = load_csv(in_csv)
    records = extract_records(rows)
    output_paths = write_excel(records, out_dir)

    elapsed = time.time() - t0
    print(f"\n{'─'*70}")
    print(f"  ✅  Done in {elapsed:.1f}s")
    print(f"  📁  Source    : {in_csv}")
    print(f"  👤  Total     : {len(records):,} employee records")
    print(f"  📊  Output    : {len(output_paths)} file(s)")
    print(f"{'─'*70}")
    for file_name, out_path, count in output_paths:
        print(f"    • {Path(out_path).name:<45} ({count} records)")
    print(f"{'─'*70}\n")
