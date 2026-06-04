"""
Scanned PDF OCR Extractor
Converts scanned PDF pages to images using PyMuPDF (no Poppler needed),
runs Tesseract OCR, and saves output as text files, CSV, and Excel.

Requirements (no external binaries except Tesseract):
    pip install pymupdf pytesseract pandas openpyxl tqdm Pillow

External dependency:
    - Tesseract OCR: https://github.com/UB-Mannheim/tesseract/wiki
      Install, then either add to PATH or set TESSERACT_CMD below.
"""

import os
import sys
import logging
from pathlib import Path

import fitz                  # PyMuPDF — no Poppler required
import pytesseract
from PIL import Image
import pandas as pd
from tqdm import tqdm
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ── Tesseract path (update if moved) ─────────────────────────────────────────
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\DidarSingh\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

# ── Config ────────────────────────────────────────────────────────────────────
OCR_LANG      = "eng"        # Tesseract language, e.g. "eng+fra"
DPI           = 300          # Rendering DPI — higher = better accuracy, slower
SAVE_TXT      = True         # One .txt file per PDF
SAVE_CSV      = True         # One .csv per PDF + combined ALL_ocr_results.csv
SAVE_EXCEL    = True         # One .xlsx per PDF + combined ALL_ocr_results.xlsx
OUTPUT_SUBDIR = "ocr_output" # Created next to the source PDF(s)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ── Core ──────────────────────────────────────────────────────────────────────

def pdf_to_images(pdf_path: Path) -> list[Image.Image]:
    """Render each PDF page to a PIL Image using PyMuPDF (no Poppler)."""
    doc    = fitz.open(str(pdf_path))
    zoom   = DPI / 72.0          # 72 dpi is PyMuPDF's default
    matrix = fitz.Matrix(zoom, zoom)
    images = []
    for page in doc:
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)
    doc.close()
    return images


def ocr_image(img: Image.Image) -> str:
    return pytesseract.image_to_string(img, lang=OCR_LANG).strip()


def extract_pdf(pdf_path: Path, out_dir: Path) -> list[dict]:
    log.info("Processing: %s", pdf_path.name)
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        images = pdf_to_images(pdf_path)
    except Exception as exc:
        log.error("Could not render %s: %s", pdf_path.name, exc)
        return []

    rows       = []
    page_texts = []

    for page_num, img in enumerate(
        tqdm(images, desc=f"  {pdf_path.name}", unit="pg", leave=False), start=1
    ):
        text = ocr_image(img)
        page_texts.append(f"=== Page {page_num} ===\n{text}\n")
        rows.append({
            "File Name":      pdf_path.name,
            "Page Number":    page_num,
            "Extracted Text": text,
        })

    if SAVE_TXT:
        txt_path = out_dir / f"{pdf_path.stem}_ocr.txt"
        txt_path.write_text("\n".join(page_texts), encoding="utf-8")
        log.info("  TXT  → %s", txt_path.name)

    return rows


# ── Writers ───────────────────────────────────────────────────────────────────

def save_csv(rows: list[dict], out_dir: Path, name: str):
    if not rows:
        return
    path = out_dir / name
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    log.info("  CSV  → %s  (%d rows)", name, len(rows))


def save_excel(rows: list[dict], out_dir: Path, name: str):
    if not rows:
        return
    path = out_dir / name
    df   = pd.DataFrame(rows)

    col_widths = {"File Name": 35, "Page Number": 12, "Extracted Text": 120}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="OCR Results")
        ws = writer.sheets["OCR Results"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[cell.row].height = 80

    log.info("  XLSX → %s  (%d rows)", name, len(rows))


def _write_outputs(rows: list[dict], out_dir: Path, stem: str):
    if SAVE_CSV:
        save_csv(rows, out_dir, f"{stem}_ocr.csv")
    if SAVE_EXCEL:
        save_excel(rows, out_dir, f"{stem}_ocr.xlsx")


# ── Runners ───────────────────────────────────────────────────────────────────

def process_file(pdf_path: str, output_dir: str | None = None):
    p        = Path(pdf_path)
    out_root = Path(output_dir) if output_dir else p.parent / OUTPUT_SUBDIR

    rows = extract_pdf(p, out_root)
    _write_outputs(rows, out_root, p.stem)
    log.info("Done → %s", out_root)


def process_folder(folder: str, output_dir: str | None = None):
    folder   = Path(folder)
    out_root = Path(output_dir) if output_dir else folder / OUTPUT_SUBDIR

    pdf_files = sorted(folder.glob("*.pdf"))
    if not pdf_files:
        log.warning("No PDF files found in: %s", folder)
        return

    log.info("Found %d PDF(s) in %s", len(pdf_files), folder)

    all_rows: list[dict] = []

    for pdf_path in pdf_files:
        rows = extract_pdf(pdf_path, out_root)
        all_rows.extend(rows)
        _write_outputs(rows, out_root, pdf_path.stem)   # per-file output

    # Combined across all files
    if SAVE_CSV:
        save_csv(all_rows, out_root, "ALL_ocr_results.csv")
    if SAVE_EXCEL:
        save_excel(all_rows, out_root, "ALL_ocr_results.xlsx")

    log.info("Done → %s", out_root)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if not args:
        process_folder(os.getcwd())
        return

    if os.path.isdir(args[0]):
        process_folder(args[0], args[1] if len(args) > 1 else None)

    elif os.path.isfile(args[0]) and args[0].lower().endswith(".pdf"):
        process_file(args[0], args[1] if len(args) > 1 else None)

    else:
        print("Usage:")
        print("  python scanned_pdf_ocr.py                          # all PDFs in current dir")
        print("  python scanned_pdf_ocr.py <folder>                 # all PDFs in folder")
        print("  python scanned_pdf_ocr.py <folder> <output_dir>    # custom output folder")
        print("  python scanned_pdf_ocr.py <file.pdf>               # single file")
        print("  python scanned_pdf_ocr.py <file.pdf> <output_dir>  # single file, custom output")
        sys.exit(1)


if __name__ == "__main__":
    main()
