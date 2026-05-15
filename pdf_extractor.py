"""
PDF to Excel Extractor
Extracts tables and text from PDF files into formatted Excel workbooks.
Output: {filename}_Extracted.xlsx

Requirements:
    pip install pdfplumber openpyxl
"""

import os
import sys
import pdfplumber
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

TABLE_HEADER_FILL  = PatternFill("solid", fgColor="2F5496")
TABLE_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
TABLE_ROW_FILL_ODD = PatternFill("solid", fgColor="DCE6F1")
TABLE_ROW_FILL_EVN = PatternFill("solid", fgColor="FFFFFF")
TEXT_FONT          = Font(name="Calibri", size=10)
PAGE_HEADER_FONT   = Font(bold=True, size=11, color="1F3864")
PAGE_HEADER_FILL   = PatternFill("solid", fgColor="D9E1F2")
THIN_BORDER_SIDE   = Side(style="thin", color="B8CCE4")
CELL_BORDER        = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def auto_fit_columns(ws, min_width=8, max_width=60):
    for col in ws.columns:
        best = min_width
        for cell in col:
            if cell.value:
                length = len(str(cell.value).split("\n")[0])
                best = max(best, min(length + 2, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


def write_table(ws, table_data, start_row):
    """Write a 2-D list as a styled table; return the next free row."""
    if not table_data:
        return start_row

    # Normalise row lengths
    max_cols = max(len(r) for r in table_data)
    rows = [r + [""] * (max_cols - len(r)) for r in table_data]

    for r_idx, row in enumerate(rows):
        is_header = r_idx == 0
        fill = (TABLE_HEADER_FILL if is_header
                else (TABLE_ROW_FILL_ODD if r_idx % 2 == 1 else TABLE_ROW_FILL_EVN))
        font = TABLE_HEADER_FONT if is_header else Font(size=10)

        for c_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=c_idx + 1,
                           value=str(value).strip() if value else "")
            cell.fill   = fill
            cell.font   = font
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return start_row + len(rows) + 1          # +1 blank separator row


def write_text_block(ws, text, start_row):
    """Write a text block; return the next free row."""
    if not text or not text.strip():
        return start_row

    for line in text.split("\n"):
        line = line.strip()
        if not line:
            start_row += 1
            continue
        cell = ws.cell(row=start_row, column=1, value=line)
        cell.font      = TEXT_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        start_row += 1

    return start_row + 1                       # blank separator row


def write_page_header(ws, page_num, current_row):
    """Write a styled 'Page N' label."""
    cell = ws.cell(row=current_row, column=1, value=f"── Page {page_num} ──")
    cell.font      = PAGE_HEADER_FONT
    cell.fill      = PAGE_HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 18
    return current_row + 2


# ── Core extractor ────────────────────────────────────────────────────────────

def extract_pdf_to_excel(pdf_path: str, output_dir: str | None = None) -> str:
    """
    Extract a single PDF to Excel.
    Returns the path of the created .xlsx file.
    """
    base_name   = os.path.splitext(os.path.basename(pdf_path))[0]
    out_name    = f"{base_name}_Extracted.xlsx"
    out_folder  = output_dir or os.path.dirname(pdf_path) or "."
    out_path    = os.path.join(out_folder, out_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted"

    # Freeze the first row so headers stay visible while scrolling
    ws.freeze_panes = "A2"

    current_row = 1

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            current_row = write_page_header(ws, page_num, current_row)

            # ── Collect tables on this page ──────────────────────────────────
            tables = page.extract_tables()

            if tables:
                # Get bounding boxes so we can skip text that overlaps tables
                table_bboxes = [t.bbox for t in page.find_tables()]
                has_bbox     = bool(table_bboxes)

                for table in tables:
                    current_row = write_table(ws, table, current_row)

                # Extract text outside table regions
                if has_bbox:
                    # Crop away table areas and get remaining text
                    remaining = page
                    for bbox in table_bboxes:
                        try:
                            remaining = remaining.outside_bbox(bbox)
                        except Exception:
                            pass
                    text = remaining.extract_text() or ""
                else:
                    text = ""

            else:
                text = page.extract_text() or ""

            if text.strip():
                current_row = write_text_block(ws, text, current_row)

    # ── Post-processing ───────────────────────────────────────────────────────
    auto_fit_columns(ws)

    # Widen column A a bit more for text-heavy content
    ws.column_dimensions["A"].width = max(
        ws.column_dimensions["A"].width, 60
    )

    wb.save(out_path)
    return out_path


# ── Batch runner ──────────────────────────────────────────────────────────────

def process_folder(folder: str, output_dir: str | None = None):
    pdf_files = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".pdf")
    ]

    if not pdf_files:
        print(f"No PDF files found in: {folder}")
        return

    print(f"Found {len(pdf_files)} PDF file(s).\n")

    for pdf_path in pdf_files:
        print(f"  Processing: {os.path.basename(pdf_path)} ...", end=" ", flush=True)
        try:
            out = extract_pdf_to_excel(pdf_path, output_dir)
            print(f"→ {os.path.basename(out)}")
        except Exception as exc:
            print(f"FAILED — {exc}")

    print("\nDone.")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if not args:
        # Default: process all PDFs in the current working directory
        target = os.getcwd()
        process_folder(target)
        return

    if os.path.isdir(args[0]):
        out_dir = args[1] if len(args) > 1 else None
        process_folder(args[0], out_dir)

    elif os.path.isfile(args[0]) and args[0].lower().endswith(".pdf"):
        out_dir = args[1] if len(args) > 1 else None
        print(f"Processing: {args[0]} ...", end=" ", flush=True)
        out = extract_pdf_to_excel(args[0], out_dir)
        print(f"→ {out}")
        print("Done.")

    else:
        print("Usage:")
        print("  python pdf_extractor.py                         # all PDFs in current dir")
        print("  python pdf_extractor.py <folder>                # all PDFs in folder")
        print("  python pdf_extractor.py <folder> <output_dir>   # custom output folder")
        print("  python pdf_extractor.py <file.pdf>              # single file")
        print("  python pdf_extractor.py <file.pdf> <output_dir> # single file, custom output")
        sys.exit(1)


if __name__ == "__main__":
    main()
