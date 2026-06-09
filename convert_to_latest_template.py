"""
Convert Excel files from old format to Latest Template format.

This script:
1. Reads files with "Standardized_Data" sheet in old format (handles both spellings)
2. Converts to Latest Template format
3. Merges duplicate entries (same Last Name + First Name + Middle Name + Suffix)
4. Fills "Data Subject Type" column with "Employee"
5. Saves the converted file
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from pathlib import Path

# Column mapping from OLD format to NEW format
COLUMN_MAPPING = {
    'DOCID': ('DOCID', 1, 1),
    'Last Name': ('Last Name', 4, 2),
    'First Name': ('First Name', 5, 3),
    'Middle Name': ('Middle Name', 6, 4),
    'Suffix': ('Suffix', 7, 5),
    'Data Subject Type': ('Data Subject Type', None, 6),  # NEW - fill with "Employee"
    'Residential Address': ('Residential Address', 8, 7),
    'State of Residence (if US)': ('State of Residence (if US)', 12, 8),
    'Country of Residence': ('Country of Residence', 11, 9),
    'City': ('City', 9, 10),
    'Province of Residence (if Canada)': ('Province of Residence (if Canada)', 13, 11),
    'Zip Code': ('Zip Code', 10, 12),
    'Address Comments': ('Address Comments', 14, 13),
    'Phone Number': ('Phone Number', 38, 14),
    'Email Address - Personal': ('Email Address - Personal', 39, 15),
    'PI Notes': ('PI Notes', 15, 16),
    'Contact Information': ('Contact Information', 18, 17),
    'Government- Issued Identification': ('Government- Issued Identification', 16, 18),
    'Social Security Number': ('Social Security Number', 35, 19),
    'Passport Number': ('Passport Number', 27, 20),
    'Passport Country': ('Passport Country', 28, 21),
    'Driver\'s License Number': ('Driver\'s License Number', 31, 22),
    'DL Issuing Country': ('DL Issuing Country', 32, 23),
    'DL Issuing Province (if Canada)': ('DL Issuing Province (if Canada)', 34, 24),
    'DL Issuing State (if US)': ('DL Issuing State (if US)', 33, 25),
    'Government-Issued ID Number': ('Government-Issued ID Number', 29, 26),
    'Government ID Issuing Country': ('Government ID Issuing Country', 30, 27),
    'Health Related Information': ('Health Related Information', 21, 28),
    'Birth Information': ('Birth Information', 17, 29),
    'Full Date of Birth (MM/DD/YYYY)': ('Full Date of Birth (MM/DD/YYYY)', 26, 30),
    'Financial Account Information': ('Financial Account Information', 19, 31),
    'Access Credentials (Non-Financial Account)': ('Access Credentials (Non-Financial Account)', 20, 32),
    'Biometric Data': ('Biometric Data', 22, 33),
    'Demographic Information': ('Demographic Information', 24, 34),
    'Family Information': ('Family Information', 23, 35),
    'Student-Related Information': ('Student-Related Information', None, 36),  # NEW
    'Work-Related Information': ('Work-Related Information', 25, 37),
    'Employee Identification Number': ('Employee Identification Number', 37, 38),
}

# New template header in order
NEW_HEADERS = [
    'DOCID', 'Last Name', 'First Name', 'Middle Name', 'Suffix',
    'Data Subject Type', 'Residential Address', 'State of Residence (if US)',
    'Country of Residence', 'City', 'Province of Residence (if Canada)',
    'Zip Code', 'Address Comments', 'Phone Number', 'Email Address - Personal',
    'PI Notes', 'Contact Information', 'Government- Issued Identification',
    'Social Security Number', 'Passport Number', 'Passport Country',
    'Driver\'s License Number', 'DL Issuing Country', 'DL Issuing Province (if Canada)',
    'DL Issuing State (if US)', 'Government-Issued ID Number', 'Government ID Issuing Country',
    'Health Related Information', 'Birth Information', 'Full Date of Birth (MM/DD/YYYY)',
    'Financial Account Information', 'Access Credentials (Non-Financial Account)',
    'Biometric Data', 'Demographic Information', 'Family Information',
    'Student-Related Information', 'Work-Related Information', 'Employee Identification Number'
]


def read_old_format(filepath):
    """Read data from Standardized_Data sheet in old format."""
    try:
        wb = openpyxl.load_workbook(filepath)

        # Find the sheet (case insensitive, handle both spellings)
        sheet_name = None
        for name in wb.sheetnames:
            name_lower = name.lower()
            # Check for both "standardized_data" and "standerdized_data" spellings
            if 'standardized_data' in name_lower or 'standerdized_data' in name_lower:
                sheet_name = name
                break

        if not sheet_name:
            print(f"  ERROR: 'Standardized_Data' sheet not found. Available sheets: {wb.sheetnames}")
            return None

        ws = wb[sheet_name]

        # Read header
        old_headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Read data rows (skip header)
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check if row has any data
            if any(cell is not None for cell in row):
                data.append(row)

        wb.close()
        return old_headers, data

    except Exception as e:
        print(f"  ERROR reading file: {str(e)}")
        return None


def convert_row(old_headers, old_row):
    """Convert a row from old format to new format."""
    # Create a dict from old headers and row for easier lookup
    old_data = {header: old_row[i] if i < len(old_row) else None
                for i, header in enumerate(old_headers) if header}

    new_row = []
    for header in NEW_HEADERS:
        if header == 'Data Subject Type':
            new_row.append('Employee')
        elif header == 'Student-Related Information':
            new_row.append(None)
        else:
            new_row.append(old_data.get(header))

    return new_row


def merge_duplicates(data, headers):
    """
    Merge rows with same Last Name + First Name + Middle Name + Suffix.
    Keep first occurrence, discard duplicates.
    """
    seen = {}
    merged = []

    # Convert headers to new format for merging
    for row in data:
        last_idx = headers.index('Last Name') if 'Last Name' in headers else None
        first_idx = headers.index('First Name') if 'First Name' in headers else None
        middle_idx = headers.index('Middle Name') if 'Middle Name' in headers else None
        suffix_idx = headers.index('Suffix') if 'Suffix' in headers else None

        if last_idx is not None and first_idx is not None:
            # Create a key from name parts
            last = row[last_idx] if last_idx < len(row) else None
            first = row[first_idx] if first_idx < len(row) else None
            middle = row[middle_idx] if middle_idx < len(row) else None
            suffix = row[suffix_idx] if suffix_idx < len(row) else None

            key = (str(last), str(first), str(middle), str(suffix))

            if key not in seen:
                seen[key] = True
                merged.append(row)

    return merged


def save_converted_file(filepath, data):
    """Save converted data to the same file or a new file."""
    try:
        # Create new workbook with new template structure
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Standerdized_Data'

        # Write headers
        for col_idx, header in enumerate(NEW_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to same file
        wb.save(filepath)
        wb.close()
        return True

    except Exception as e:
        print(f"  ERROR saving file: {str(e)}")
        return False


def process_file(filepath):
    """Process a single file."""
    print(f"\nProcessing: {filepath}")

    # Read old format
    result = read_old_format(filepath)
    if result is None:
        return False

    old_headers, old_data = result
    print(f"  Found {len(old_data)} data rows")

    # Convert each row
    converted_data = []
    for row in old_data:
        new_row = convert_row(old_headers, row)
        converted_data.append(new_row)

    print(f"  Converted {len(converted_data)} rows")

    # Merge duplicates
    merged_data = merge_duplicates(converted_data, NEW_HEADERS)
    print(f"  After merging duplicates: {len(merged_data)} rows")

    # Save
    if save_converted_file(filepath, merged_data):
        print(f"  SUCCESS: File saved")
        return True
    else:
        return False


def main():
    """Main entry point."""
    if len(sys.argv) > 1:
        # Process specific file
        filepath = sys.argv[1]
        if os.path.exists(filepath):
            process_file(filepath)
        else:
            print(f"File not found: {filepath}")
    else:
        # Process all .xlsx files in current directory
        print("Searching for .xlsx files with 'Standerdized_Data' sheet...\n")

        files_found = 0
        for file in Path('.').glob('*.xlsx'):
            if not str(file).startswith('~'):  # Skip temp files
                files_found += 1
                process_file(str(file))

        if files_found == 0:
            print("No .xlsx files found in current directory.")
            print("\nUsage:")
            print("  python convert_to_latest_template.py")
            print("    - Processes all .xlsx files in current directory")
            print("  python convert_to_latest_template.py <filepath>")
            print("    - Processes specific file")


if __name__ == '__main__':
    main()
