"""
New Hire Extractor → Excel
Processes PDFs directly — text-based and scanned (OCR via Tesseract).
Extracts: Name (Last, First) and Position ID into new_hire_extracted.xlsx.

Setup:
    pip install pymupdf pytesseract Pillow openpyxl tqdm
    Install Tesseract: https://github.com/UB-Mannheim/tesseract/wiki

Run:
    python new_hire_extract.py                       # prompts for folder
    python new_hire_extract.py "C:\\path\\to\\pdfs"  # pass folder directly
"""

import os
import re
import sys
import fitz
import pytesseract
import openpyxl
from PIL import Image
from tqdm import tqdm
from openpyxl.styles import Font, PatternFill, Alignment

# ── Tesseract path ─────────────────────────────────────────────────────────────
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\DidarSingh\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
)

DPI = 300
# Minimum characters on a page to consider it "has real text" (not scanned)
MIN_TEXT_CHARS = 30


# ── Regex patterns ─────────────────────────────────────────────────────────────

# Name: "LAST, FIRST" or "LAST, FIRST MIDDLE" — all-caps format on HR forms
RE_NAME = re.compile(
    r"\b([A-Z][A-Z'\-]{1,}),\s+([A-Z][A-Z'\-]+(?:\s+[A-Z][A-Z'\-]+)*)\b"
)

# Position ID: common labels on new hire forms
RE_POSITION_ID = re.compile(
    r"(?:Position\s*(?:ID|No\.?|Number|#)?|Pos\.?\s*(?:ID|No\.?)?|"
    r"Job\s*(?:ID|Code|No\.?)|Requisition\s*(?:ID|No\.?)|Req\.?\s*(?:ID|No\.?)?)[\s:\-#]*"
    r"([A-Z0-9]{3,20})\b",
    re.IGNORECASE,
)


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract full text from a PDF, using OCR for scanned pages."""
    doc = fitz.open(pdf_path)
    zoom = DPI / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    pages_text = []

    with tqdm(total=len(doc), unit="page", ncols=70, leave=False,
              bar_format="  {l_bar}{bar}| {n_fmt}/{total_fmt} pages") as pbar:
        for page in doc:
            native = page.get_text().strip()

            if len(native) >= MIN_TEXT_CHARS:
                # Text-based page — use native extraction
                pages_text.append(native)
            else:
                # Scanned page — OCR via Tesseract
                pix = page.get_pixmap(matrix=matrix, alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                pages_text.append(pytesseract.image_to_string(img).strip())

            pbar.update(1)

    doc.close()
    return "\n".join(pages_text)


def extract_fields(text: str) -> dict:
    result = {"last_name": "", "first_name": "", "position_id": ""}

    m = RE_NAME.search(text)
    if m:
        result["last_name"]  = m.group(1).strip()
        result["first_name"] = m.group(2).strip()

    m = RE_POSITION_ID.search(text)
    if m:
        result["position_id"] = m.group(1).strip()

    return result


def build_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Hires"

    headers = ["PDF File", "Last Name", "First Name", "Position ID"]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["pdf_file"])
        ws.cell(row=row_idx, column=2, value=row["last_name"])
        ws.cell(row=row_idx, column=3, value=row["first_name"])
        ws.cell(row=row_idx, column=4, value=row["position_id"])

    for col, width in zip([1, 2, 3, 4], [30, 18, 22, 16]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1].strip()
    else:
        print("=" * 60)
        print("  New Hire Extractor → Excel")
        print("=" * 60)
        folder = input("\nEnter folder path containing PDF files:\n> ").strip().strip('"')

    if not os.path.isdir(folder):
        print(f"\nError: '{folder}' is not a valid folder.")
        input("\nPress Enter to exit...")
        sys.exit(1)

    pdf_files = sorted(f for f in os.listdir(folder) if f.lower().endswith(".pdf"))

    if not pdf_files:
        print(f"\nNo PDF files found in:\n  {folder}")
        input("\nPress Enter to exit...")
        sys.exit(0)

    print(f"\nFolder : {folder}")
    print(f"PDFs   : {len(pdf_files)} file(s) found")
    print("-" * 60)

    rows = []
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"\n[{i}/{len(pdf_files)}] {pdf_file}")
        try:
            text   = extract_text_from_pdf(os.path.join(folder, pdf_file))
            fields = extract_fields(text)
            fields["pdf_file"] = pdf_file
            rows.append(fields)

            name = f"{fields['last_name']}, {fields['first_name']}".strip(", ")
            pos  = fields["position_id"] or "—"
            print(f"  Name: {name or '?'}  |  Position ID: {pos}")

        except Exception as e:
            print(f"  ERROR: {e}")
            rows.append({"pdf_file": pdf_file, "last_name": "", "first_name": "", "position_id": "ERROR"})

    output_path = os.path.join(folder, "new_hire_extracted.xlsx")
    build_excel(rows, output_path)

    print("\n" + "=" * 60)
    print(f"  Done!  {len(rows)} record(s) written.")
    print(f"  Output : {output_path}")
    print("=" * 60)
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
