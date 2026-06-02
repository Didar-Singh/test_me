"""
Scanned PDF OCR Extractor
Converts scanned PDF pages to images, runs Tesseract OCR, and saves output
as plain text files and/or a structured CSV/Excel workbook.

Requirements:
    pip install pytesseract pdf2image pandas openpyxl tqdm Pillow

External dependency (must be installed separately):
    - Tesseract OCR:  https://github.com/UB-Mannheim/tesseract/wiki  (Windows installer)
    - Poppler:        https://github.com/oschwartz10612/poppler-windows/releases  (Windows)
      After installing, add both bin directories to PATH, or set the paths below.
"""

import os
import sys
import csv
import logging
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import pandas as pd
from tqdm import tqdm

# ── Optional: set explicit paths if not on PATH ───────────────────────────────
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# POPPLER_PATH = r"C:\poppler\Library\bin"   # set to None if Poppler is on PATH
POPPLER_PATH = None

# ── Config ────────────────────────────────────────────────────────────────────
OCR_LANG        = "eng"          # Tesseract language code(s), e.g. "eng+fra"
DPI             = 300            # Higher = better accuracy, slower
SAVE_TXT        = True           # Write one .txt file per PDF
SAVE_CSV        = True           # Write combined CSV
SAVE_EXCEL      = True           # Write combined Excel workbook
OUTPUT_SUBDIR   = "ocr_output"   # Sub-folder created next to the input PDFs

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ── Core functions ────────────────────────────────────────────────────────────

def ocr_page(image: Image.Image) -> str:
    """Run Tesseract on a single PIL image and return extracted text."""
    return pytesseract.image_to_string(image, lang=OCR_LANG)


def extract_pdf(pdf_path: str, output_dir: str) -> list[dict]:
    """
    OCR all pages of a scanned PDF.
    Returns a list of row-dicts: {file, page, text}.
    """
    pdf_path   = Path(pdf_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log.info("Opening: %s", pdf_path.name)

    convert_kwargs = {"dpi": DPI, "fmt": "jpeg"}
    if POPPLER_PATH:
        convert_kwargs["poppler_path"] = POPPLER_PATH

    try:
        pages = convert_from_path(str(pdf_path), **convert_kwargs)
    except Exception as exc:
        log.error("Failed to convert %s: %s", pdf_path.name, exc)
        return []

    rows          = []
    page_texts    = []

    for page_num, img in enumerate(
        tqdm(pages, desc=f"  {pdf_path.name}", unit="page", leave=False), start=1
    ):
        text = ocr_page(img).strip()
        page_texts.append(f"=== Page {page_num} ===\n{text}\n")
        rows.append({
            "File Name":       pdf_path.name,
            "Page Number":     page_num,
            "Extracted Text":  text,
        })

    if SAVE_TXT:
        txt_path = output_dir / f"{pdf_path.stem}_ocr.txt"
        txt_path.write_text("\n".join(page_texts), encoding="utf-8")
        log.info("  Saved text → %s", txt_path.name)

    return rows


# ── Output writers ─────────────────────────────────────────────────────────────

def save_csv(rows: list[dict], output_dir: Path, filename: str = "ocr_results.csv"):
    if not rows:
        return
    path = output_dir / filename
    df   = pd.DataFrame(rows)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    log.info("CSV saved → %s  (%d rows)", path, len(df))


def save_excel(rows: list[dict], output_dir: Path, filename: str = "ocr_results.xlsx"):
    if not rows:
        return
    path = output_dir / filename
    df   = pd.DataFrame(rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="OCR Results")
        ws = writer.sheets["OCR Results"]

        # Auto-fit columns (cap at 120 for the text column)
        col_widths = {"File Name": 35, "Page Number": 12, "Extracted Text": 120}
        for col_idx, col_name in enumerate(df.columns, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
                col_name, 20
            )

        # Wrap text in the extracted-text column and set row height
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = __import__("openpyxl").styles.Alignment(
                    wrap_text=True, vertical="top"
                )
                ws.row_dimensions[cell.row].height = 80

    log.info("Excel saved → %s  (%d rows)", path, len(df))


# ── Batch runner ───────────────────────────────────────────────────────────────

def process_folder(folder: str, output_dir: str | None = None):
    folder     = Path(folder)
    out_root   = Path(output_dir) if output_dir else folder / OUTPUT_SUBDIR
    out_root.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(folder.glob("*.pdf"))
    if not pdf_files:
        log.warning("No PDF files found in: %s", folder)
        return

    log.info("Found %d PDF file(s) in %s", len(pdf_files), folder)

    all_rows: list[dict] = []

    for pdf_path in pdf_files:
        rows = extract_pdf(str(pdf_path), str(out_root))
        all_rows.extend(rows)

        # Individual output named after each source file
        if SAVE_CSV:
            save_csv(rows, out_root, f"{pdf_path.stem}_ocr.csv")
        if SAVE_EXCEL:
            save_excel(rows, out_root, f"{pdf_path.stem}_ocr.xlsx")

    # Combined summary across all files
    if SAVE_CSV:
        save_csv(all_rows, out_root, "ALL_ocr_results.csv")
    if SAVE_EXCEL:
        save_excel(all_rows, out_root, "ALL_ocr_results.xlsx")

    log.info("Done. Output folder: %s", out_root)


def process_file(pdf_path: str, output_dir: str | None = None):
    pdf_path = Path(pdf_path)
    out_root = Path(output_dir) if output_dir else pdf_path.parent / OUTPUT_SUBDIR
    out_root.mkdir(parents=True, exist_ok=True)

    rows = extract_pdf(str(pdf_path), str(out_root))

    if SAVE_CSV:
        save_csv(rows, out_root, f"{pdf_path.stem}_ocr.csv")
    if SAVE_EXCEL:
        save_excel(rows, out_root, f"{pdf_path.stem}_ocr.xlsx")

    log.info("Done. Output folder: %s", out_root)


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if not args:
        process_folder(os.getcwd())
        return

    if os.path.isdir(args[0]):
        out_dir = args[1] if len(args) > 1 else None
        process_folder(args[0], out_dir)

    elif os.path.isfile(args[0]) and args[0].lower().endswith(".pdf"):
        out_dir = args[1] if len(args) > 1 else None
        process_file(args[0], out_dir)

    else:
        print("Usage:")
        print("  python scanned_pdf_ocr.py                          # all PDFs in current dir")
        print("  python scanned_pdf_ocr.py <folder>                 # all PDFs in folder")
        print("  python scanned_pdf_ocr.py <folder> <output_dir>    # custom output folder")
        print("  python scanned_pdf_ocr.py <file.pdf>               # single file")
        print("  python scanned_pdf_ocr.py <file.pdf> <output_dir>  # single file, custom output")
        sys.exit(1)


if __name__ == "__main__":
    main()
