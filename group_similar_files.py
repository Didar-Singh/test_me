#!/usr/bin/env python3
"""
Similar File Grouper
Scans .txt files, detects form type / field structure, groups similar
files into named sets, and writes an Excel report.

Usage:
  python group_similar_files.py                  # scan current folder
  python group_similar_files.py path/to/folder   # scan given folder
  python group_similar_files.py a.txt b.txt ...  # specific files

Output: similar_files_report.xlsx  (File Name | Set Name | Form Type | Matched Fields)

Requirements:
  pip install pandas openpyxl
"""

import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies.  Run:  pip install pandas openpyxl")
    sys.exit(1)


# ── Known form signatures ────────────────────────────────────────────────────
# Each entry: (set_name, [keyword_patterns]).
# A file matches the FIRST type whose majority of keywords are present.

KNOWN_FORMS = [
    (
        "New Hire Form",
        [r"NEW HIRE", r"Date of Hire", r"Badge No", r"SSN Verified", r"Harassment Prevention"],
    ),
    (
        "Payroll Change Notice",
        [r"PAYROLL CHANGE", r"Reason for Change", r"Current Rate", r"Future Rate", r"FLSA"],
    ),
    (
        "W-2 / Tax Form",
        [r"W-2", r"Wages.*Tips", r"Federal Income Tax", r"Social Security Wages", r"Medicare"],
    ),
    (
        "I-9 Employment Eligibility",
        [r"I-9", r"Employment Eligibility", r"List A", r"List B", r"List C", r"Document Title"],
    ),
    (
        "Performance Review",
        [r"Performance Review", r"Rating", r"Goals", r"Competencies", r"Review Period"],
    ),
    (
        "Expense Report",
        [r"Expense Report", r"Reimbursement", r"Receipt", r"Business Purpose", r"Total Amount"],
    ),
]

# Generic field-group signatures used to cluster unknown forms
FIELD_GROUP_SIGNATURES = {
    "Employee Profile":   [r"Employee Name", r"Date of Birth", r"Social Security", r"Gender"],
    "Contact Details":    [r"Street Address", r"City.*State", r"Telephone", r"Email"],
    "Job Information":    [r"Job Title", r"Department", r"Office Location", r"Reports To"],
    "Compensation":       [r"Annual Salary", r"Hourly Rate", r"Pay Cycle", r"Car Allowance"],
    "Change Record":      [r"Reason for Change", r"Current Rate", r"Future Rate", r"Effective Date"],
    "Leave / Absence":    [r"Leave of Absence", r"LOA", r"Return Date", r"Leave Type"],
    "Benefits":           [r"Health Insurance", r"Dental", r"401[Kk]", r"Beneficiary"],
    "Structured Table":   [],   # fallback for CSV/TSV files detected separately
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def count_keyword_hits(text: str, patterns: list) -> int:
    return sum(1 for p in patterns if re.search(p, text, re.IGNORECASE))


def detect_known_type(text: str) -> str | None:
    """Return the first known form type whose majority of keywords match."""
    best_type, best_score = None, 0
    for form_name, patterns in KNOWN_FORMS:
        hits = count_keyword_hits(text, patterns)
        ratio = hits / len(patterns) if patterns else 0
        if ratio > 0.4 and hits > best_score:
            best_type, best_score = form_name, hits
    return best_type


def detect_generic_signature(text: str) -> str | None:
    """For unknown forms, return the dominant field-group signature."""
    best_group, best_hits = None, 0
    for group_name, patterns in FIELD_GROUP_SIGNATURES.items():
        if not patterns:
            continue
        hits = count_keyword_hits(text, patterns)
        if hits > best_hits:
            best_group, best_hits = group_name, hits
    return best_group if best_hits >= 2 else None


def extract_matched_fields(text: str) -> list[str]:
    """Return a short list of recognised field names found in the file."""
    known_fields = [
        "Employee Name", "Date of Hire", "Job Title", "Department",
        "Office Location", "Annual Salary", "Hourly Rate", "Pay Cycle",
        "Reason for Change", "Current Rate", "Future Rate", "File #",
        "File No", "Badge No", "Email Address", "Date of Birth",
        "Social Security Number", "Marital Status", "Employee Type",
    ]
    return [f for f in known_fields if re.search(rf"\b{re.escape(f)}\b", text, re.IGNORECASE)]


def is_csv_like(text: str) -> bool:
    """True if the first non-empty line looks like a CSV/TSV header."""
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        for sep in (",", "\t", "|"):
            parts = line.split(sep)
            if len(parts) >= 4:
                return True
        break
    return False


def analyse_file(path: Path) -> dict:
    text = path.read_text(encoding="utf-8", errors="ignore")

    if is_csv_like(text):
        set_name = "Structured Table (CSV/TSV)"
        form_type = "Structured Table"
    else:
        form_type = detect_known_type(text)
        if form_type:
            set_name = form_type
        else:
            generic = detect_generic_signature(text)
            set_name = f"Similar: {generic}" if generic else "Uncategorised"
            form_type = set_name

    matched = extract_matched_fields(text)

    return {
        "File Name":      path.name,
        "File Path":      str(path),
        "Set Name":       set_name,
        "Form Type":      form_type,
        "Matched Fields": ", ".join(matched) if matched else "—",
        "Field Count":    len(matched),
    }


# ── File collection ──────────────────────────────────────────────────────────

def collect_files(args: list) -> list[Path]:
    if not args:
        return sorted(Path(".").glob("*.txt"))
    paths = []
    for arg in args:
        p = Path(arg)
        if p.is_dir():
            paths.extend(sorted(p.glob("*.txt")))
        elif p.is_file() and p.suffix.lower() == ".txt":
            paths.append(p)
        else:
            print(f"WARNING: not found or not a .txt file, skipping: {arg}")
    return paths


# ── Excel output ─────────────────────────────────────────────────────────────

SET_COLOURS = [
    "D9E1F2", "E2EFDA", "FCE4D6", "FFF2CC", "DDEBF7",
    "EAD1DC", "D0E4F1", "F0F0F0", "E8F5E9", "FFF9C4",
]

def write_excel(rows: list[dict], out_path: Path):
    df = pd.DataFrame(rows, columns=[
        "File Name", "Set Name", "Form Type", "Matched Fields", "Field Count", "File Path"
    ])
    df.sort_values(["Set Name", "File Name"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="File Groups")
        ws = writer.sheets["File Groups"]

        # Header style
        header_fill = PatternFill("solid", fgColor="2F5597")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ws.row_dimensions[1].height = 22

        # Colour bands per set
        set_names = df["Set Name"].unique().tolist()
        colour_map = {s: SET_COLOURS[i % len(SET_COLOURS)] for i, s in enumerate(set_names)}

        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2   # 1-indexed + header
            fill = PatternFill("solid", fgColor=colour_map[row["Set Name"]])
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Column widths
        col_widths = {"File Name": 36, "Set Name": 30, "Form Type": 28,
                      "Matched Fields": 55, "Field Count": 14, "File Path": 60}
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

        # Summary sheet
        summary_data = (
            df.groupby("Set Name")
            .agg(File_Count=("File Name", "count"), Avg_Fields=("Field Count", "mean"))
            .reset_index()
            .rename(columns={"File_Count": "File Count", "Avg_Fields": "Avg Matched Fields"})
        )
        summary_data["Avg Matched Fields"] = summary_data["Avg Matched Fields"].round(1)
        summary_data.to_excel(writer, index=False, sheet_name="Summary")

        ws2 = writer.sheets["Summary"]
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 30


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    cli_args = [a for a in sys.argv[1:] if a not in ("-h", "--help")]

    if "-h" in sys.argv or "--help" in sys.argv:
        print(__doc__)
        return

    files = collect_files(cli_args)
    if not files:
        print("No .txt files found.")
        return

    print(f"Analysing {len(files)} file(s)...")

    rows = []
    for path in files:
        try:
            result = analyse_file(path)
            rows.append(result)
            print(f"  {result['File Name']:40s}  →  {result['Set Name']}")
        except Exception as e:
            print(f"  ERROR {path.name}: {e}")

    if not rows:
        print("Nothing to write.")
        return

    out_path = Path("similar_files_report.xlsx")
    write_excel(rows, out_path)

    # Console summary
    sets = defaultdict(list)
    for r in rows:
        sets[r["Set Name"]].append(r["File Name"])

    print(f"\nSaved: {out_path}  ({len(rows)} files, {len(sets)} set(s))\n")
    print(f"{'Set Name':<35}  Files")
    print("-" * 55)
    for set_name, file_list in sorted(sets.items()):
        print(f"  {set_name:<33}  {len(file_list):>3}  ({', '.join(file_list[:3])}{'...' if len(file_list) > 3 else ''})")


if __name__ == "__main__":
    main()
