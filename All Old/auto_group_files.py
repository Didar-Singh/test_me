#!/usr/bin/env python3
"""
Auto File Typer & Grouper
Dynamically analyses .txt files, identifies data nature from content,
groups similar files into named sets, and exports an Excel report.

NO hardcoded form types — fully data-driven.

Usage:
  python auto_group_files.py                     # scan current folder
  python auto_group_files.py path/to/folder      # scan given folder
  python auto_group_files.py -r path/to/folder   # recursive scan
  python auto_group_files.py --threshold 0.25    # looser grouping (default 0.30)
  python auto_group_files.py --debug             # print detected fields per file

Output: file_groups_report.xlsx

Requirements:
  pip install pandas openpyxl
"""

import re
import sys
from pathlib import Path
from collections import Counter, defaultdict

try:
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies.  Run:  pip install pandas openpyxl")
    sys.exit(1)


# ── Constants ────────────────────────────────────────────────────────────────

STOP_WORDS = {
    'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
    'of', 'with', 'by', 'is', 'are', 'was', 'were', 'be', 'been', 'have',
    'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should',
    'may', 'might', 'shall', 'can', 'not', 'no', 'it', 'its', 'this',
    'that', 'these', 'those', 'i', 'you', 'he', 'she', 'we', 'they',
    'from', 'as', 'if', 'so', 'all', 'each', 'any', 'more', 'other',
    'please', 'per', 'n', 'a', 'b', 'c', 'x', 'yes', 'no',
}

SET_COLOURS = [
    "D9E1F2", "E2EFDA", "FCE4D6", "FFF2CC", "DDEBF7",
    "EAD1DC", "D0E4F1", "F4CCCC", "E8F5E9", "FFF9C4",
    "F3E5F5", "E8EAF6", "F1F8E9", "FFF8E1", "E0F2F1",
    "FBE9E7", "E3F2FD", "F9FBE7", "FCE4D6", "E8F4FD",
]


# ── File analysis ────────────────────────────────────────────────────────────

def detect_structure(text):
    """Return: 'form' | 'csv' | 'tsv' | 'list' | 'plain'"""
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return "empty"

    first = lines[0]

    # Tab-separated (TSV) — check first line
    if '\t' in first:
        cols = first.split('\t')
        if len(cols) >= 3:
            return "tsv"

    # Comma-separated with multiple columns
    if ',' in first and first.count(',') >= 3:
        parts = first.split(',')
        if all(len(p.strip()) < 80 for p in parts[:8]):
            return "csv"

    # Pipe-delimited
    if first.count('|') >= 3:
        return "csv"

    # Count "Label: value" lines in first 50 non-empty lines
    sample = lines[:50]
    colon_matches = sum(
        1 for l in sample
        if re.match(r'^[A-Za-z][A-Za-z0-9 /\-#&\'\.]{0,45}\s*[:\=]\s*\S', l)
    )
    if colon_matches >= 3 or (len(sample) > 0 and colon_matches / len(sample) >= 0.2):
        return "form"

    # Bullet / numbered list
    bullet = sum(1 for l in sample if re.match(r'^\s*[-•*►◆]\s+\S|^\s*\d+[.)]\s+\S', l))
    if len(sample) > 0 and bullet / len(sample) >= 0.35:
        return "list"

    return "plain"


def extract_title(text):
    """Detect a document title from the first meaningful lines."""
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines[:15]:
        if len(line) > 80 or ':' in line:
            continue
        # All-caps short phrase → clear title
        if line.isupper() and 3 <= len(line.split()) <= 10:
            return line.title()
        # Title-case phrase of reasonable length
        words = line.split()
        if 2 <= len(words) <= 8 and line[0].isupper():
            # At least half the words are capitalised
            caps = sum(1 for w in words if w and w[0].isupper())
            if caps / len(words) >= 0.6:
                return line
    return None


def extract_form_labels(text):
    """Extract field label names from form-structured text (before : or =)."""
    labels = set()

    # Pattern 1 — "Field Name: value" (single or multi-word label)
    for m in re.finditer(
        r'^([A-Za-z][A-Za-z0-9 /\-#&\'\.]{0,44}?)\s*[:\=]\s*\S',
        text, re.MULTILINE
    ):
        raw = m.group(1).strip()
        if _is_good_label(raw):
            labels.add(raw.lower())

    # Pattern 2 — standalone ALL-CAPS label on its own line
    for m in re.finditer(r'^([A-Z][A-Z &\-\/]{2,40})\s*$', text, re.MULTILINE):
        raw = m.group(1).strip()
        if 2 <= len(raw.split()) <= 6:
            labels.add(raw.lower())

    # Pattern 3 — "Label -  value" (dash separator)
    for m in re.finditer(
        r'^([A-Za-z][A-Za-z0-9 /\-#&\'\.]{1,44}?)\s+[-–]\s+\S',
        text, re.MULTILINE
    ):
        raw = m.group(1).strip()
        if _is_good_label(raw):
            labels.add(raw.lower())

    return labels


def _is_good_label(raw):
    """Filter out noise from regex matches."""
    words = raw.split()
    if not (1 <= len(words) <= 7):
        return False
    if any(c.isdigit() for c in raw[:2]):   # starts with a number
        return False
    if len(raw) < 2:
        return False
    # Reject lines that look like sentences (all stop words)
    content_words = [w for w in words if w.lower() not in STOP_WORDS]
    return len(content_words) >= 1


def extract_csv_columns(text, structure):
    """Extract column names from CSV/TSV first header line."""
    first = next((l.strip() for l in text.splitlines() if l.strip()), "")
    sep = '\t' if structure == 'tsv' else ('|' if '|' in first else ',')
    cols = {c.strip().lower() for c in first.split(sep) if c.strip() and len(c.strip()) < 60}
    return cols


def extract_keywords(text, top_n=40):
    """Top meaningful words for plain / list text."""
    words = re.findall(r'\b[a-zA-Z]{3,}\b', text.lower())
    counts = Counter(w for w in words if w not in STOP_WORDS)
    return {w for w, _ in counts.most_common(top_n)}


def analyse_file(path):
    text = path.read_text(encoding='utf-8', errors='ignore')
    structure = detect_structure(text)
    title = extract_title(text)

    if structure == 'form':
        features = extract_form_labels(text)
        feature_type = 'fields'
    elif structure in ('csv', 'tsv'):
        features = extract_csv_columns(text, structure)
        feature_type = 'columns'
    else:
        features = extract_keywords(text)
        feature_type = 'keywords'

    return {
        'path':         path,
        'name':         path.name,
        'structure':    structure,
        'title':        title,
        'features':     features,
        'feature_type': feature_type,
        'feature_count': len(features),
    }


# ── Similarity & clustering ──────────────────────────────────────────────────

def jaccard(a, b):
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def cluster_files(file_data, threshold):
    """
    Graph-based clustering via BFS on connected components.
    Two files are connected when jaccard(features) >= threshold.
    Files of different feature_type can still connect via shared tokens.
    """
    n = len(file_data)
    adj = defaultdict(set)

    for i in range(n):
        for j in range(i + 1, n):
            sim = jaccard(file_data[i]['features'], file_data[j]['features'])
            if sim >= threshold:
                adj[i].add(j)
                adj[j].add(i)

    visited = [False] * n
    clusters = []
    for start in range(n):
        if visited[start]:
            continue
        cluster = []
        queue = [start]
        while queue:
            node = queue.pop()
            if visited[node]:
                continue
            visited[node] = True
            cluster.append(node)
            for nb in adj[node]:
                if not visited[nb]:
                    queue.append(nb)
        clusters.append(sorted(cluster))

    return clusters


def name_cluster(files):
    """Generate a human-readable set name from the cluster's shared features."""
    # 1. Majority title wins
    titles = [f['title'] for f in files if f['title']]
    if titles:
        top, count = Counter(titles).most_common(1)[0]
        if count >= max(1, len(files) // 2):
            return _clean_name(top)

    # 2. Shared field labels / column names across most files
    n = len(files)
    combined = Counter()
    for f in files:
        combined.update(f['features'])

    # Pick labels present in at least half the files
    shared = [token for token, cnt in combined.most_common(30)
              if cnt >= max(1, n // 2) and len(token) >= 3
              and token not in STOP_WORDS]

    if shared:
        # Prefer longer, more descriptive tokens
        top3 = sorted(shared[:10], key=len, reverse=True)[:3]
        return ' / '.join(t.title() for t in top3)

    # 3. Fallback: structure type
    structs = Counter(f['structure'] for f in files)
    return structs.most_common(1)[0][0].title() + " Document"


def _clean_name(name):
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:60] if len(name) > 60 else name


# ── Excel output ─────────────────────────────────────────────────────────────

def write_excel(rows, out_path):
    df = pd.DataFrame(rows)
    df.sort_values(['Set Name', 'File Name'], inplace=True)
    df.reset_index(drop=True, inplace=True)

    export_cols = [
        'File Name', 'Set Name', 'Structure Type',
        'Key Fields Detected', 'Feature Count', 'File Path'
    ]

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df[export_cols].to_excel(writer, index=False, sheet_name='File Groups')
        ws = writer.sheets['File Groups']

        hdr_fill = PatternFill('solid', fgColor='2F5597')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        thin = Side(style='thin', color='CCCCCC')
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bdr
        ws.row_dimensions[1].height = 22

        set_names = df['Set Name'].unique().tolist()
        colour_map = {s: SET_COLOURS[i % len(SET_COLOURS)] for i, s in enumerate(set_names)}

        for row_idx in range(len(df)):
            excel_row = row_idx + 2
            fill = PatternFill('solid', fgColor=colour_map[df.iloc[row_idx]['Set Name']])
            for col_idx in range(1, len(export_cols) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = fill
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')

        widths = {
            'File Name': 38, 'Set Name': 35, 'Structure Type': 18,
            'Key Fields Detected': 58, 'Feature Count': 16, 'File Path': 65
        }
        for i, col in enumerate(export_cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 20)

        # --- Summary sheet ---
        summary = (
            df.groupby('Set Name')
            .agg(
                Files=('File Name', 'count'),
                Structure=('Structure Type', lambda x: x.mode()[0]),
                Avg_Fields=('Feature Count', 'mean')
            )
            .reset_index()
            .rename(columns={'Avg_Fields': 'Avg Fields'})
        )
        summary['Avg Fields'] = summary['Avg Fields'].round(1)
        summary.to_excel(writer, index=False, sheet_name='Summary')
        ws2 = writer.sheets['Summary']
        for cell in ws2[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 30


# ── CLI helpers ──────────────────────────────────────────────────────────────

def collect_files(args, recursive):
    if not args:
        pattern = '**/*.txt' if recursive else '*.txt'
        return sorted(Path('.').glob(pattern))

    paths = []
    for arg in args:
        p = Path(arg)
        if p.is_dir():
            pattern = '**/*.txt' if recursive else '*.txt'
            paths.extend(sorted(p.glob(pattern)))
        elif p.is_file() and p.suffix.lower() == '.txt':
            paths.append(p)
        else:
            print(f"WARNING: skipping (not found): {arg}")
    return paths


def parse_args(argv):
    recursive  = '-r' in argv or '--recursive' in argv
    debug      = '--debug' in argv
    threshold  = 0.30
    clean_args = []

    i = 0
    while i < len(argv):
        arg = argv[i]
        if arg in ('-r', '--recursive', '--debug'):
            pass
        elif arg == '--threshold' and i + 1 < len(argv):
            try:
                threshold = float(argv[i + 1])
            except ValueError:
                pass
            i += 1
        else:
            clean_args.append(arg)
        i += 1

    return clean_args, recursive, debug, threshold


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    raw_args = [a for a in sys.argv[1:] if a not in ('-h', '--help')]
    if '-h' in sys.argv or '--help' in sys.argv:
        print(__doc__)
        return

    clean_args, recursive, debug, threshold = parse_args(raw_args)

    files = collect_files(clean_args, recursive)
    if not files:
        print("No .txt files found.")
        return

    print(f"Analysing {len(files)} file(s)  (similarity threshold = {threshold})\n")

    file_data = []
    for path in files:
        try:
            fd = analyse_file(path)
            file_data.append(fd)
            if debug:
                feat_preview = ', '.join(sorted(fd['features'])[:5])
                print(f"  [{fd['structure']:5s}] {fd['name']:40s}  title={fd['title'] or '—'}")
                print(f"         fields({fd['feature_count']}): {feat_preview}")
            else:
                print(f"  [{fd['structure']:5s}] {fd['name']}")
        except Exception as e:
            print(f"  ERROR {path.name}: {e}")

    if not file_data:
        print("Nothing to process.")
        return

    print(f"\nClustering {len(file_data)} files...")
    clusters = cluster_files(file_data, threshold)
    print(f"  → {len(clusters)} group(s) found\n")

    rows = []
    used_names = Counter()
    for cluster_idx, indices in enumerate(clusters):
        cluster_fds = [file_data[i] for i in indices]
        base_name   = name_cluster(cluster_fds)
        used_names[base_name] += 1
        set_name = base_name if used_names[base_name] == 1 else f"{base_name} ({used_names[base_name]})"

        for fd in cluster_fds:
            # Show longest labels first (more descriptive)
            top_labels = sorted(fd['features'], key=len, reverse=True)[:7]
            rows.append({
                'File Name':          fd['name'],
                'Set Name':           set_name,
                'Structure Type':     fd['structure'],
                'Key Fields Detected': ', '.join(top_labels),
                'Feature Count':      fd['feature_count'],
                'File Path':          str(fd['path']),
            })

    out_path = Path('file_groups_report.xlsx')
    write_excel(rows, out_path)

    # Console summary
    print(f"{'Set Name':<42}  {'Files':>5}  Structure")
    print('-' * 65)
    set_map = defaultdict(list)
    struct_map = {}
    for r in rows:
        set_map[r['Set Name']].append(r['File Name'])
        struct_map[r['Set Name']] = r['Structure Type']
    for sn in sorted(set_map):
        fnames   = set_map[sn]
        preview  = ', '.join(fnames[:3]) + ('...' if len(fnames) > 3 else '')
        print(f"  {sn:<40}  {len(fnames):>5}  [{struct_map[sn]}]  ({preview})")

    print(f"\nSaved: {out_path}  ({len(rows)} files, {len(set_map)} groups)")
    print("\nTip: if groups are too broad → raise --threshold (e.g. 0.45)")
    print("     if too many singletons → lower --threshold (e.g. 0.20)")


if __name__ == '__main__':
    main()
