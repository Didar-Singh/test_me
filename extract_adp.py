"""
ADP Master Control Report - Extractor
======================================
Layout: 3 columns per employee record
  LEFT: PERSONNEL  |  MIDDLE: PAY  |  RIGHT: TAX STATUS

Multiple employees stacked vertically per page.
Employee records separated by blank lines / new name blocks.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 COMMANDS & EXAMPLES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

 1. Basic — process entire PDF, auto-name output:
       python extract_adp.py myfile.pdf

 2. Custom output filename:
       python extract_adp.py myfile.pdf results.xlsx

 3. Process specific page range (e.g. pages 2 to 5):
       python extract_adp.py myfile.pdf --pages 2-5

 4. Process specific page range + custom output:
       python extract_adp.py myfile.pdf results.xlsx --pages 2-5

 5. Single page test (e.g. just page 3):
       python extract_adp.py myfile.pdf --pages 3-3

 6. Debug mode — prints raw text per column per employee:
       python extract_adp.py myfile.pdf --debug

 7. Debug on a page range (best for troubleshooting):
       python extract_adp.py myfile.pdf --pages 1-2 --debug

 8. Install required libraries (run once):
       pip install pdfplumber openpyxl tqdm

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 FLAGS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  --pages FROM-TO   Only process pages FROM through TO (1-based)
  --debug           Print raw extracted text per employee block
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import re, sys
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from tqdm import tqdm

DEBUG = "--debug" in sys.argv

# Parse --pages FROM-TO flag
PAGE_FROM = None
PAGE_TO   = None
for arg in sys.argv[1:]:
    m = re.match(r"--pages\s*[=:]?\s*(\d+)[-:](\d+)$", arg)
    if m:
        PAGE_FROM = int(m.group(1))
        PAGE_TO   = int(m.group(2))
        break
    m2 = re.match(r"--pages\s*[=:]?\s*(\d+)$", arg)
    if m2:
        PAGE_FROM = PAGE_TO = int(m2.group(1))
        break

args = [a for a in sys.argv[1:] if not a.startswith("--")]

# ── helpers ───────────────────────────────────────────────────────────────────

def clean(s):
    return " ".join(str(s).split()).strip() if s else ""

def grep(pattern, text, group=1, default="", flags=re.IGNORECASE):
    m = re.search(pattern, text, flags)
    return clean(m.group(group)) if m else default

def mgrep(patterns, text, default=""):
    for p in (patterns if isinstance(patterns, list) else [patterns]):
        v = grep(p, text, default="")
        if v: return v
    return default

def crop_text(page, x0, y0, x1, y1):
    try:
        return page.crop((x0, y0, x1, y1)).extract_text(x_tolerance=3, y_tolerance=3) or ""
    except:
        return ""

# ── find employee record Y boundaries on page ─────────────────────────────────

def find_employee_boundaries(page):
    """
    Each employee starts with their name in the PERSONNEL column (left ~35% of page).
    Name line pattern: ALL CAPS, may have comma (LAST,FIRST or LAST FIRST).
    We detect these lines by scanning words in the left column only.
    """
    pw = page.width
    ph = page.height

    # Only look at left 35% of page for name detection
    left_col_x1 = pw * 0.35

    words = page.extract_words(x_tolerance=3, y_tolerance=3)

    # Find lines that look like employee names in left column
    # Group words by their vertical position (top coordinate, rounded to 2px)
    from collections import defaultdict
    lines_by_y = defaultdict(list)
    for w in words:
        if w["x0"] < left_col_x1:
            y_key = round(w["top"] / 2) * 2
            lines_by_y[y_key].append(w)

    # Header/footer keywords to skip
    SKIP_WORDS = {
        "PERSONNEL","PAY","TAX","STATUS","FILE","DEPT","CLOCK","SSN","TITLE",
        "SEX","RACE","OCCUP","DATE","HIRE","BIRTH","DATES","GROSS","SALARY",
        "MARITAL","FEDERAL","STATE","RATE","LWW","NWW","CONTINUED","PAGE",
        "AUTOPAY","MASTER","CONTROL","TOTALSOURCE","COMPANY","CODE","ADP",
        "EXEMPTIONS","BI-WKLY","RATE","CALC","STATUS","ACTIVE","TERM",
        "SET","FOR","PURGE","(CONTINUED)"
    }

    name_ys = []
    for y_key in sorted(lines_by_y.keys()):
        line_words = sorted(lines_by_y[y_key], key=lambda w: w["x0"])
        line_text  = " ".join(w["text"] for w in line_words).strip()

        # Skip very short or purely numeric lines
        if len(line_text) < 3 or re.match(r"^[\d\s\.\-/]+$", line_text):
            continue

        # Skip known header words
        upper_words = set(line_text.upper().split())
        if upper_words & SKIP_WORDS and not re.search(r"[,]", line_text):
            continue

        # Skip lines containing field labels like "File:", "Dept:", etc.
        if re.search(r"\b(File|Dept|Clock|SSN|Title|Sex|Race|Occup|Hire|Birth|Date|Status|LWW|NWW)\s*[:\d]", line_text, re.I):
            continue

        # Name characteristics:
        # - Mostly uppercase letters
        # - May contain comma (LAST,FIRST)
        # - 2+ words or one word >=4 chars
        # - Starts with a capital letter
        # - No numbers (except maybe address — but address comes AFTER name)
        if re.match(r"^[A-Z][A-Z\s,\.\'\-]+$", line_text.upper()):
            # Must be at least one real name-looking token
            tokens = re.split(r"[,\s]+", line_text)
            if any(len(t) >= 2 for t in tokens):
                actual_y = line_words[0]["top"]
                name_ys.append(actual_y)

    if DEBUG:
        print(f"  [page w={pw:.0f} h={ph:.0f}] Name Y positions: {[f'{y:.1f}' for y in name_ys]}")

    if not name_ys:
        return []

    # Build vertical slices
    boundaries = []
    name_ys = sorted(name_ys)
    for i, y in enumerate(name_ys):
        y0 = max(0, y - 3)
        y1 = (name_ys[i+1] - 3) if i + 1 < len(name_ys) else ph
        boundaries.append((y0, y1))

    return boundaries

# ── parse one employee record ─────────────────────────────────────────────────

def parse_record(page, y0, y1, page_num):
    pw = page.width

    # ADP Master Control 3-column split:
    # PERSONNEL: 0 – ~35%
    # PAY:      ~35% – ~60%
    # TAX:      ~60% – 100%
    personnel_txt = crop_text(page, 0,        y0, pw*0.36, y1)
    pay_txt       = crop_text(page, pw*0.36,  y0, pw*0.60, y1)
    tax_txt       = crop_text(page, pw*0.60,  y0, pw,      y1)

    if DEBUG:
        print(f"\n{'─'*60}")
        print(f"  PAGE {page_num}  y={y0:.0f}-{y1:.0f}")
        print(f"  [PERSONNEL]\n{personnel_txt}")
        print(f"  [PAY]\n{pay_txt}")
        print(f"  [TAX]\n{tax_txt}")

    rec = {"_page": page_num}
    p_lines = [l.strip() for l in personnel_txt.splitlines() if l.strip()]

    # ── NAME ──────────────────────────────────────────────────────────────────
    SKIP = {"PERSONNEL","PAY","TAX","STATUS","FILE","DEPT","CLOCK","SSN",
            "TITLE","SEX","RACE","OCCUP","DATES","DATE","HIRE","BIRTH",
            "GROSS","SALARY","MARITAL","FEDERAL","(CONTINUED)","CONTINUED"}
    name_raw = ""
    for ln in p_lines[:5]:
        toks = set(ln.upper().split(",")[0].split())
        if toks & SKIP:
            continue
        if re.match(r"^[A-Z][A-Z\s,\.\'\-]+$", ln.upper()) and len(ln) >= 3:
            name_raw = ln.strip()
            break

    if name_raw:
        if "," in name_raw:
            parts = name_raw.split(",", 1)
            rec["Last Name"]  = clean(parts[0])
            rec["First Name"] = clean(parts[1])
        else:
            toks = name_raw.split()
            rec["Last Name"]  = " ".join(toks[:-1]) if len(toks) > 1 else name_raw
            rec["First Name"] = toks[-1] if len(toks) > 1 else ""
    else:
        rec["Last Name"] = rec["First Name"] = ""

    # ── ADDRESS (present on some records) ─────────────────────────────────────
    # Address appears right after name: street line, then city,state zip
    addr1 = addr2 = city = state = zipcode = ""
    street_m = re.search(
        r"(?:^|\n)(\d+\s+[A-Z0-9\s]+(?:RD|ST|AVE|DR|LN|WAY|BLVD|CT|PL|TERR?|HWY|PIKE|ROAD|STREET|AVENUE|DRIVE|LANE)\b[^\n]*)",
        personnel_txt, re.IGNORECASE | re.MULTILINE)
    if street_m:
        addr1 = clean(street_m.group(1))
        # city,state zip usually on next line
        after = personnel_txt[street_m.end():]
        csz_m = re.search(r"([A-Z\s]+),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)", after, re.IGNORECASE)
        if csz_m:
            city    = clean(csz_m.group(1))
            state   = clean(csz_m.group(2))
            zipcode = clean(csz_m.group(3))
        else:
            # Try "CITY ST ZIP" without comma
            csz_m2 = re.search(r"([A-Z\s]{2,})\s+([A-Z]{2})\s+(\d{5})", after, re.IGNORECASE)
            if csz_m2:
                city    = clean(csz_m2.group(1))
                state   = clean(csz_m2.group(2))
                zipcode = clean(csz_m2.group(3))

    rec["Address"]  = addr1
    rec["City"]     = city
    rec["State"]    = state
    rec["Zip"]      = zipcode

    # ── PERSONNEL FIELDS ──────────────────────────────────────────────────────
    rec["File #"]   = mgrep([r"File[:\s]+(\d+)", r"File\s*[:#]\s*(\w+)"], personnel_txt)
    rec["Dept"]     = grep(r"Dept[:\s]+(\S+)", personnel_txt)
    rec["Clock"]    = grep(r"Clock[:\s]+(\S+)", personnel_txt)
    rec["SSN"]      = grep(r"SSN[:\s]+(\S+)", personnel_txt)
    rec["Title"]    = grep(r"Title[:\s]+(\S+)", personnel_txt)
    rec["Sex"]      = grep(r"Sex[:\s]+(\w)", personnel_txt)
    rec["Race"]     = grep(r"Race[:\s]+(\w+)", personnel_txt)
    rec["Occup"]    = grep(r"Occup[:\s]+(\w+)", personnel_txt)
    rec["Status"]   = grep(r"Status[:\s]+(\w+)", personnel_txt)
    rec["Set for Purge"] = "Yes" if re.search(r"Set\s+for\s+Purge", personnel_txt, re.I) else ""

    # ── DATES ─────────────────────────────────────────────────────────────────
    rec["Date 1"]    = grep(r"Date\s*1[:\s]+([\d/]+)", personnel_txt)
    rec["Date 3"]    = grep(r"Date\s*3[:\s]+([\d/]+)", personnel_txt)
    rec["Hire Date"] = mgrep([r"Hire[:\s]+([\d/]+)", r"Hire\s+Date[:\s]+([\d/]+)"], personnel_txt)
    rec["Birth Date"]= mgrep([r"Birth[:\s]+([\d/]+)", r"DOB[:\s]+([\d/]+)"], personnel_txt)
    rec["Date 9"]    = grep(r"Date\s*9[:\s]+([\d/]+)", personnel_txt)

    # ── PAY FIELDS ────────────────────────────────────────────────────────────
    rec["Gross"]     = mgrep([r"Gross[:\s]+([\d,\.]+)", r"Gross\s+Pay[:\s]+([\d,\.]+)"], pay_txt)
    rec["Salary"]    = grep(r"Salary[:\s]+([\d,\.]+)", pay_txt)
    rec["Bi-Wkly"]   = grep(r"Bi-Wkly[:\s]+([\d,\.]+)", pay_txt)
    rec["Rate Calc"] = grep(r"Rate\s*Calc[:\s]+(\w+)", pay_txt)
    rec["LWW"]       = grep(r"LWW[:\s]+(\d+)", pay_txt)
    rec["NWW"]       = grep(r"NWW[:\s]+(\d+)", pay_txt)

    # ── TAX STATUS FIELDS ─────────────────────────────────────────────────────
    rec["Marital Status"] = mgrep([
        r"Marital\s+Status[:\s]+(\S+)",
        r"M-(\w+)",
        r"S-(\w+)",
    ], tax_txt)
    # Full marital line e.g. "Marital Status: S-SINGLE" or "Marital Status: M-MARRIED"
    ms = grep(r"Marital\s+Status[:\s]*([^\n]+)", tax_txt)
    if ms:
        rec["Marital Status"] = ms

    rec["Federal Exemptions"] = mgrep([
        r"Federal[:\s]+(\d+)\s*Exemptions?",
        r"(\d+)\s*Exemptions?\s*Federal",
        r"Federal\s*\n\s*(\d+)\s*Exemptions?",
    ], tax_txt)

    # State lines: "59 PA SUIDI", "401D DOYLESTOWN B", "09 PA" etc.
    state_lines = re.findall(r"(\d{2,4}[A-Z0-9\s]+(?:SUIDI?|LOCAL|SUI|SDI|[A-Z]{2})?\b[^\n]*)", tax_txt)
    rec["State Tax Lines"] = "; ".join(clean(s) for s in state_lines[:4])

    # Individual state fields
    rec["State Code 1"] = state_lines[0].strip() if len(state_lines) > 0 else ""
    rec["State Code 2"] = state_lines[1].strip() if len(state_lines) > 1 else ""
    rec["State Code 3"] = state_lines[2].strip() if len(state_lines) > 2 else ""

    rec["Location"]     = grep(r"([A-Z]{3,}\s+[A-Z])\s*$", tax_txt)  # e.g. "WARWICK T"

    return rec

# ── process full PDF ──────────────────────────────────────────────────────────

def extract_all(pdf_path, page_from=None, page_to=None):
    employees = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)

        # Resolve page range (1-based input -> 0-based index)
        p_from = max(1, page_from) if page_from else 1
        p_to   = min(total, page_to) if page_to else total

        if p_from > total or p_from > p_to:
            print(f"\n  ❌ Invalid page range {p_from}-{p_to} (PDF has {total} pages)")
            return []

        page_indices = range(p_from - 1, p_to)   # convert to 0-based
        page_count   = len(page_indices)

        print(f"  Total pages in PDF : {total}")
        if page_from or page_to:
            print(f"  Processing pages   : {p_from} to {p_to}  ({page_count} page(s))")
        else:
            print(f"  Processing pages   : all {total}")

        with tqdm(total=page_count, desc="📄 Reading pages",
                  unit="pg", colour="cyan", ncols=65) as pbar:
            for page_num in page_indices:
                page   = pdf.pages[page_num]
                bounds = find_employee_boundaries(page)

                if DEBUG:
                    print(f"\n══ PAGE {page_num+1} — {len(bounds)} employee(s) found ══")

                if not bounds:
                    if DEBUG:
                        print(f"  Raw page text sample:")
                        txt = page.extract_text() or ""
                        print(txt[:600])
                    pbar.update(1)
                    continue

                for y0, y1 in bounds:
                    try:
                        rec = parse_record(page, y0, y1, page_num + 1)
                        if rec.get("Last Name") or rec.get("File #"):
                            employees.append(rec)
                    except Exception as e:
                        tqdm.write(f"  ⚠️  Page {page_num+1} y={y0:.0f}: {e}")
                pbar.update(1)

    return employees

# ── Excel output ──────────────────────────────────────────────────────────────

COLUMNS = [
    # Name
    "Last Name", "First Name",
    # Address
    "Address", "City", "State", "Zip",
    # Personnel
    "File #", "Dept", "Clock", "SSN", "Title",
    "Status", "Sex", "Race", "Occup", "Set for Purge",
    # Dates
    "Date 1", "Date 3", "Hire Date", "Birth Date", "Date 9",
    # Pay
    "Gross", "Salary", "Bi-Wkly", "Rate Calc", "LWW", "NWW",
    # Tax
    "Marital Status", "Federal Exemptions",
    "State Code 1", "State Code 2", "State Code 3", "State Tax Lines",
    "Location",
    # Meta
    "_page",
]

# Color per section
SEC_COLOR = {
    "Last Name":"1F4E79","First Name":"1F4E79",
    "Address":"1F3864","City":"1F3864","State":"1F3864","Zip":"1F3864",
    "File #":"375623","Dept":"375623","Clock":"375623","SSN":"375623",
    "Title":"375623","Status":"375623","Sex":"375623","Race":"375623",
    "Occup":"375623","Set for Purge":"375623",
    "Date 1":"7B2C2C","Date 3":"7B2C2C","Hire Date":"7B2C2C",
    "Birth Date":"7B2C2C","Date 9":"7B2C2C",
    "Gross":"7B4A00","Salary":"7B4A00","Bi-Wkly":"7B4A00",
    "Rate Calc":"7B4A00","LWW":"7B4A00","NWW":"7B4A00",
    "Marital Status":"4A235A","Federal Exemptions":"4A235A",
    "State Code 1":"4A235A","State Code 2":"4A235A","State Code 3":"4A235A",
    "State Tax Lines":"4A235A","Location":"4A235A",
    "_page":"555555",
}

SEC_LABELS = {
    "Last Name":       "👤 Name",
    "Address":         "📍 Address",
    "File #":          "🗂 Personnel",
    "Date 1":          "📅 Dates",
    "Gross":           "💰 Pay",
    "Marital Status":  "🧾 Tax Status",
    "_page":           "ℹ Meta",
}

def write_excel(employees, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ADP Master Control"

    BD  = Side(style="thin", color="BFBFBF")
    BDR = Border(left=BD, right=BD, top=BD, bottom=BD)
    NF  = Font(name="Arial", size=10)
    AF  = PatternFill("solid", fgColor="EEF2F7")

    # Row 1: Section group headers (merged cells)
    current_label = None
    span_start = 1
    for ci, col in enumerate(COLUMNS, 1):
        lbl = SEC_LABELS.get(col)
        if lbl and lbl != current_label:
            if current_label is not None:
                ws.merge_cells(start_row=1, start_column=span_start,
                               end_row=1,   end_column=ci-1)
                c = ws.cell(row=1, column=span_start, value=current_label)
                color = SEC_COLOR.get(COLUMNS[span_start-1], "333333")
                c.fill = PatternFill("solid", fgColor=color)
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = BDR
            current_label = lbl
            span_start = ci
    # Last section
    if current_label:
        ws.merge_cells(start_row=1, start_column=span_start,
                       end_row=1,   end_column=len(COLUMNS))
        c = ws.cell(row=1, column=span_start, value=current_label)
        color = SEC_COLOR.get(COLUMNS[span_start-1], "333333")
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BDR
    ws.row_dimensions[1].height = 22

    # Row 2: Column headers
    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.fill = PatternFill("solid", fgColor=SEC_COLOR.get(col, "1F4E79"))
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BDR
    ws.row_dimensions[2].height = 28

    # Data rows
    with tqdm(total=len(employees), desc="💾 Writing Excel",
              unit="row", colour="yellow", ncols=65) as pbar:
        for ri, emp in enumerate(employees, 3):
            fill = AF if ri % 2 == 0 else None
            for ci, col in enumerate(COLUMNS, 1):
                c = ws.cell(row=ri, column=ci, value=emp.get(col, ""))
                c.font = NF; c.border = BDR
                c.alignment = Alignment(vertical="center")
                if fill: c.fill = fill
            pbar.update(1)

    # Column widths
    W = {"Last Name":16,"First Name":16,"Address":26,"City":16,"State":6,
         "Zip":9,"File #":10,"Dept":9,"Clock":9,"SSN":14,"Title":12,
         "Status":9,"Sex":5,"Race":7,"Occup":7,"Set for Purge":12,
         "Date 1":11,"Date 3":11,"Hire Date":11,"Birth Date":11,"Date 9":11,
         "Gross":11,"Salary":11,"Bi-Wkly":10,"Rate Calc":9,"LWW":7,"NWW":7,
         "Marital Status":16,"Federal Exemptions":14,
         "State Code 1":18,"State Code 2":18,"State Code 3":18,
         "State Tax Lines":32,"Location":14,"_page":6}
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = W.get(col, 12)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"
    wb.save(out_path)

# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(args) < 1:
        print(__doc__)
        sys.exit(1)

    pdf_path = args[0]
    out_path = args[1] if len(args) > 1 else \
               str(Path(pdf_path).stem + "_employees.xlsx")

    if not Path(pdf_path).exists():
        print(f"\n❌ File not found: {pdf_path}")
        sys.exit(1)

    # Build flags summary for display
    flags = []
    if PAGE_FROM or PAGE_TO:
        flags.append(f"pages {PAGE_FROM}-{PAGE_TO}")
    if DEBUG:
        flags.append("DEBUG")
    flags_str = "  " + " | ".join(flags) if flags else ""

    print(f"\n{'='*55}")
    print(f"  ADP Master Control Extractor")
    print(f"{'='*55}")
    print(f"  Input : {pdf_path}")
    print(f"  Output: {out_path}")
    if flags_str:
        print(f"  Flags :{flags_str}")
    print(f"{'='*55}")

    employees = extract_all(pdf_path, page_from=PAGE_FROM, page_to=PAGE_TO)
    print(f"\n  👥 Records found: {len(employees)}")

    if not employees:
        print("\n  ❌ No records extracted.")
        print("  Tips:")
        print(f"    • Debug mode  : python extract_adp.py \"{pdf_path}\" --debug")
        print(f"    • Single page : python extract_adp.py \"{pdf_path}\" --pages 1-1 --debug")
        print(f"    • Check range : python extract_adp.py \"{pdf_path}\" --pages 1-3\n")
        sys.exit(1)

    write_excel(employees, out_path)

    print(f"\n{'='*55}")
    print(f"  ✅ Done! → {out_path}")
    print(f"{'='*55}\n")
