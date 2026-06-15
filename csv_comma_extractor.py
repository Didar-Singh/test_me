"""
CSV Comma Line Extractor
------------------------
Searches one or more CSV/TXT files for lines containing a comma
(optionally filtered by a keyword), extracts N lines before and after
each match, and writes results to Excel in VERTICAL or HORIZONTAL layout.

Usage:
    python csv_comma_extractor.py [options]

Options:
    --files       One or more input file paths (default: all .csv/.txt in current dir)
    --search      Keyword to filter comma-lines (case-insensitive). Leave blank = ALL comma lines.
    --before      Number of context lines before the match (default: 2)
    --after       Number of context lines after the match  (default: 2)
    --layout      'vertical' or 'horizontal' (default: horizontal)
    --output      Output Excel file path (default: output_extracted.xlsx)

Examples:
    python csv_comma_extractor.py --files data.csv --search DIDAR --layout horizontal
    python csv_comma_extractor.py --files file1.csv file2.txt --layout vertical
"""

import argparse
import re
import sys
import time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required.  Install with:  pip install openpyxl")


# ── Colours ────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")
CTX_B_FILL = PatternFill("solid", fgColor="DEEAF1")
CTX_A_FILL = PatternFill("solid", fgColor="FFF2CC")
SEP_FILL   = PatternFill("solid", fgColor="F2F2F2")
GREY_FILL  = PatternFill("solid", fgColor="EDEDED")
BODY_FONT  = Font(name="Calibri", size=11)
BOLD_FONT  = Font(name="Calibri", size=11, bold=True)


# ── Progress bar ───────────────────────────────────────────────
def progress_bar(current: int, total: int, label: str = "", width: int = 38):
    pct    = current / total if total else 1
    filled = int(width * pct)
    bar    = "█" * filled + "░" * (width - filled)
    print(f"\r  [{bar}] {int(pct*100):3d}%  {label:<28}", end="", flush=True)
    if current >= total:
        print()


# ── Helpers ────────────────────────────────────────────────────
def split_line(line: str) -> list[str]:
    parts = re.split(r",|\s{2,}", line)
    return [p.strip() for p in parts if p.strip()]


def find_matches(lines: list[str], search: str) -> list[int]:
    keyword = search.upper()
    return [
        i for i, line in enumerate(lines)
        if "," in line and (not keyword or keyword in line.upper())
    ]


def read_file(fp: Path) -> list[str]:
    text = fp.read_text(encoding="utf-8", errors="replace")
    return [l.rstrip("\r\n") for l in text.splitlines()]


# ── VERTICAL writer ────────────────────────────────────────────
def write_vertical(all_blocks: list, output_path: Path):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Extracted Lines"

    max_cols = max(
        (len(split_line(ln)) for blk in all_blocks for _, ln, _ in blk["rows"]),
        default=0,
    )
    headers = ["File", "Line #", "Role", "Raw Line"] + [f"Col {i+1}" for i in range(max_cols)]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    total = len(all_blocks)
    print("\n  Writing rows to Excel...")
    first = True
    for idx, blk in enumerate(all_blocks, 1):
        progress_bar(idx, total, f"block {idx}/{total}")
        if not first:
            ws.append(["── ── ──"] + [""] * (len(headers) - 1))
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = SEP_FILL
                ws.cell(r, c).font = Font(name="Calibri", size=10, italic=True, color="999999")
        first = False

        mi = blk["match_idx_abs"]
        for line_num, raw, is_match in blk["rows"]:
            pos = line_num - 1
            if pos < mi:
                role, fill = f"Context before ({mi - pos})", CTX_B_FILL
            elif pos == mi:
                role, fill = "Match", MATCH_FILL
            else:
                role, fill = f"Context after ({pos - mi})", CTX_A_FILL
            cols     = split_line(raw)
            row_data = [blk["file"], line_num, role, raw] + \
                       [cols[i] if i < len(cols) else "" for i in range(max_cols)]
            ws.append(row_data)
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill
                ws.cell(r, c).font = BOLD_FONT if is_match else BODY_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 60
    for i in range(max_cols):
        ws.column_dimensions[get_column_letter(5 + i)].width = 16
    ws.freeze_panes = "A2"

    print("  Saving file...")
    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}  ({ws.max_row - 1} rows)\n")


# ── HORIZONTAL writer ──────────────────────────────────────────
def write_horizontal(all_blocks: list, before: int, after: int, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Lines"

    labels, fills_hdr = ["File", "Match Line #"], [HDR_FILL, HDR_FILL]
    for i in range(before, 0, -1):
        labels    += [f"Before {i} – Line #", f"Before {i} – Raw Line"]
        fills_hdr += [CTX_B_FILL, CTX_B_FILL]
    labels    += ["Match – Line #", "Match – Raw Line"]
    fills_hdr += [MATCH_FILL, MATCH_FILL]
    for i in range(1, after + 1):
        labels    += [f"After {i} – Line #", f"After {i} – Raw Line"]
        fills_hdr += [CTX_A_FILL, CTX_A_FILL]

    ws.append(labels)
    for cell, fill in zip(ws[1], fills_hdr):
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_fills = (
        [CTX_B_FILL] * (before * 2) +
        [MATCH_FILL] * 2 +
        [CTX_A_FILL] * (after * 2)
    )

    total = len(all_blocks)
    print("\n  Writing rows to Excel...")
    for idx, blk in enumerate(all_blocks, 1):
        progress_bar(idx, total, f"match {idx}/{total}")
        match_ln = blk["rows"][blk["match_idx"]][0]
        data = [blk["file"], match_ln]
        for ln, raw, _ in blk["rows"]:
            data += [ln, raw]
        ws.append(data)
        r = ws.max_row
        for c in [1, 2]:
            ws.cell(r, c).fill = GREY_FILL
            ws.cell(r, c).font = BOLD_FONT
        for offset, fill in enumerate(row_fills, start=3):
            cell       = ws.cell(r, offset)
            cell.fill  = fill
            cell.font  = BOLD_FONT if fill == MATCH_FILL else BODY_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 14
    col = 3
    for _ in range(before + 1 + after):
        ws.column_dimensions[get_column_letter(col)].width     = 10
        ws.column_dimensions[get_column_letter(col + 1)].width = 55
        col += 2
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    print("  Saving file...")
    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}  ({ws.max_row - 1} match rows)\n")


# ── Main ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Extract comma-containing lines + context to Excel.")
    parser.add_argument("--files",  nargs="*")
    parser.add_argument("--search", default="")
    parser.add_argument("--before", type=int, default=2)
    parser.add_argument("--after",  type=int, default=2)
    parser.add_argument("--layout", choices=["vertical", "horizontal"], default="horizontal")
    parser.add_argument("--output", default="output_extracted.xlsx")
    args = parser.parse_args()

    if args.files:
        file_paths = [Path(f) for f in args.files]
        missing = [f for f in file_paths if not f.exists()]
        if missing:
            sys.exit(f"Files not found: {', '.join(str(f) for f in missing)}")
    else:
        file_paths = sorted(Path(".").glob("*.csv")) + sorted(Path(".").glob("*.txt"))
        if not file_paths:
            sys.exit("No .csv/.txt files found. Use --files to specify paths.")

    print("=" * 65)
    print("  CSV Comma Line Extractor")
    print("=" * 65)
    print(f"  Files   : {len(file_paths)}")
    print(f"  Search  : {args.search or '(any comma line)'}")
    print(f"  Context : {args.before} before  +  {args.after} after")
    print(f"  Layout  : {args.layout}")
    print(f"  Output  : {args.output}")
    print("=" * 65)

    all_blocks    = []
    total_matches = 0
    total_files   = len(file_paths)

    print("\n  Scanning files...")
    for f_idx, fp in enumerate(file_paths, 1):
        progress_bar(f_idx, total_files, fp.name)
        try:
            lines = read_file(fp)
        except Exception as e:
            print(f"\n  Warning: {fp.name}: {e}")
            continue

        match_indices  = find_matches(lines, args.search)
        total_matches += len(match_indices)

        for mi in match_indices:
            start     = max(0, mi - args.before)
            end       = min(len(lines) - 1, mi + args.after)
            rows      = [(j + 1, lines[j], j == mi) for j in range(start, end + 1)]
            local_idx = mi - start
            all_blocks.append({
                "file":          fp.name,
                "rows":          rows,
                "match_idx":     local_idx,
                "match_idx_abs": mi,
            })

    print(f"\n  Total matches found: {total_matches}")

    if not all_blocks:
        print("  No matches. Try a different --search term or leave it blank.")
        return

    if args.layout == "horizontal":
        write_horizontal(all_blocks, args.before, args.after, Path(args.output))
    else:
        write_vertical(all_blocks, Path(args.output))

    print("=" * 65)
    print("  Done!")
    print("=" * 65)


if __name__ == "__main__":
    main()
